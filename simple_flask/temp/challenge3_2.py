from openpyxl import Workbook
import openpyxl

wb = openpyxl.load_workbook('courses.xlsx')
wb.create_sheet('combine')
wb2013 = Workbook()
wb2014 = Workbook()
wb2015 = Workbook()
wb2016 = Workbook()

wb2013['Sheet'].title = '2013'
wb2014['Sheet'].title = '2014'
wb2015['Sheet'].title = '2015'
wb2016['Sheet'].title = '2016'

for row in wb['students']:
	for row2 in wb['time']:
		if row[1].value == row2[1].value:
			wb['combine']['A'+str(row[1].row)] = row[0].value
			wb['combine']['B'+str(row[1].row)] = row[1].value
			wb['combine']['C'+str(row[1].row)] = row[2].value
			wb['combine']['D'+str(row[1].row)] = row2[2].value

wb2013['2013']['A1'] = wb2014['2014']['A1'] =wb2015['2015']['A1'] =wb2016['2016']['A1'] = wb['combine']['A1'].value
wb2013['2013']['B1'] = wb2014['2014']['B1'] =wb2015['2015']['B1'] =wb2016['2016']['B1'] = wb['combine']['B1'].value
wb2013['2013']['C1'] = wb2014['2014']['C1'] =wb2015['2015']['C1'] =wb2016['2016']['C1'] = wb['combine']['C1'].value
wb2013['2013']['D1'] = wb2014['2014']['D1'] =wb2015['2015']['D1'] =wb2016['2016']['D1'] = wb['combine']['D1'].value

for row in wb['combine']:
	if str(row[0].value).split('-')[0] == '2013':
		wb2013['2013'].append([row[0].value,row[1].value,row[2].value,row[3].value])
	elif str(row[0].value).split('-')[0] == '2014':
		wb2014['2014'].append([row[0].value,row[1].value,row[2].value,row[3].value])
	elif str(row[0].value).split('-')[0] == '2015':
		wb2015['2015'].append([row[0].value,row[1].value,row[2].value,row[3].value])
	elif str(row[0].value).split('-')[0] == '2016':
		wb2016['2016'].append([row[0].value,row[1].value,row[2].value,row[3].value])
wb.save('courses.xlsx')
wb2013.save('2013.xlsx')
wb2014.save('2014.xlsx')
wb2015.save('2015.xlsx')
wb2016.save('2016.xlsx')